import streamlit as st
import ee
import geemap.foliumap as geemap
import geopandas as gpd
import pandas as pd
import plotly.express as px
import plotly.io as pio
import tempfile
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

# Configuração da página
st.set_page_config(layout='wide')

# Inicializar Earth Engine
ee.Initialize()

# === FUNÇÕES ===

@st.cache_data
def buscar_imagem_sentinel(_area_fc):
    colecao = ee.ImageCollection("COPERNICUS/S2_SR_HARMONIZED") \
        .filterBounds(_area_fc.geometry()) \
        .filterDate('2021-01-01', '2025-12-31') \
        .filter(ee.Filter.lt('CLOUDY_PIXEL_PERCENTAGE', 10)) \
        .sort('system:time_start', False)
    return colecao.first()

def calcular_ndvi(imagem):
    return imagem.normalizedDifference(['B8', 'B4']).rename('NDVI')

def ler_geojson(uploaded_file):
    if uploaded_file is not None:
        gdf = gpd.read_file(uploaded_file)
        geojson = gdf.__geo_interface__
        return ee.FeatureCollection(geojson)
    return None

def estatisticas_ndvi(ndvi, area_fc):
    stats = ndvi.reduceRegion(
        reducer=ee.Reducer.minMax().combine(
            reducer2=ee.Reducer.mean(),
            sharedInputs=True
        ),
        geometry=area_fc.geometry(),
        scale=10,
        maxPixels=1e13
    )
    return stats.getInfo()

def valor_seguro(dado, casas=4):
    return round(dado, casas) if dado is not None else None

def gerar_excel_estatisticas(df, fig_bar, fig_pie):
    wb = Workbook()
    ws = wb.active
    ws.title = "NDVI Estatísticas"

    # Preencher tabela de dados
    ws.append(df.columns.tolist())
    for row in df.itertuples(index=False):
        ws.append(list(row))

    # Salvar figura do gráfico de barras em arquivo temporário
    fd_bar, path_bar = tempfile.mkstemp(suffix='.png')
    os.close(fd_bar)
    pio.write_image(fig_bar, path_bar, format='png', scale=2)

    # Salvar figura do gráfico de pizza em arquivo temporário
    fd_pie, path_pie = tempfile.mkstemp(suffix='.png')
    os.close(fd_pie)
    pio.write_image(fig_pie, path_pie, format='png', scale=2)

    # Adicionar gráfico de barras na planilha
    ws_bar = wb.create_sheet(title="Gráfico Barras")
    img_bar = XLImage(path_bar)
    img_bar.width, img_bar.height = 480, 360
    ws_bar.add_image(img_bar, "A1")

    # Adicionar gráfico de pizza na planilha
    ws_pie = wb.create_sheet(title="Gráfico Pizza")
    img_pie = XLImage(path_pie)
    img_pie.width, img_pie.height = 480, 360
    ws_pie.add_image(img_pie, "A1")

    # Salvar arquivo Excel temporário e limpar imagens temporárias
    fd_excel, path_excel = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd_excel)
    wb.save(path_excel)
    os.remove(path_bar)
    os.remove(path_pie)

    return path_excel

# === INTERFACE ===

st.title("🛰️ Análise de NDVI com Sentinel-2")

st.markdown("""
### Como usar esta ferramenta

1. Na aba **Dados**, envie um arquivo GeoJSON com a área que deseja analisar.
2. Vá para a aba **Estatísticas** para processar e visualizar os valores de NDVI mínimo, máximo e médio.
3. Utilize os gráficos gerados para interpretar os dados.
4. Opcionalmente, baixe um relatório em Excel com os dados e gráficos.
5. Na aba **Mapa**, visualize a distribuição espacial do NDVI sobre sua área de estudo.

**Fonte dos dados:** [Sentinel-2 Surface Reflectance (COPERNICUS/S2_SR_HARMONIZED)](https://developers.google.com/earth-engine/datasets/catalog/COPERNICUS_S2_SR_HARMONIZED)
""")

tab_dados, tab_estatisticas, tab_mapa = st.tabs(["Dados", "Estatísticas", "Mapa"])

with tab_dados:
    uploaded_file = st.file_uploader("🌎 Envie o arquivo GeoJSON com sua área de interesse", type="geojson")
    if uploaded_file:
        st.success("Arquivo carregado com sucesso. Vá para a aba Estatísticas.")

with tab_estatisticas:
    if uploaded_file:
        with st.spinner("Processando imagem e calculando NDVI..."):
            area_fc = ler_geojson(uploaded_file)
            imagem = buscar_imagem_sentinel(area_fc)
            ndvi = calcular_ndvi(imagem)
            stats = estatisticas_ndvi(ndvi, area_fc)

            ndvi_min = stats.get('NDVI_min')
            ndvi_max = stats.get('NDVI_max')
            ndvi_mean = stats.get('NDVI_mean')

            if None in [ndvi_min, ndvi_max, ndvi_mean]:
                st.warning("Não foi possível calcular NDVI para esta geometria. Verifique se a área possui dados válidos (ex: sem nuvens, água ou fora da cobertura do Sentinel-2).")
            else:
                df_stats = pd.DataFrame([{
                    'NDVI mínimo': valor_seguro(ndvi_min),
                    'NDVI máximo': valor_seguro(ndvi_max),
                    'NDVI médio': valor_seguro(ndvi_mean)
                }])

                st.subheader("Estatísticas de NDVI")
                st.dataframe(df_stats)

                cores_ndvi = {
                    "NDVI mínimo": "#d73027",
                    "NDVI máximo": "#1a9850",
                    "NDVI médio":  "#fee08b",
                }

                # Gráfico de barras
                df_melted = df_stats.melt(var_name="Indicador", value_name="Valor")
                fig_ndvi_bar = px.bar(
                    df_melted,
                    x="Indicador",
                    y="Valor",
                    color="Indicador",
                    color_discrete_map=cores_ndvi,
                    title="Resumo do NDVI"
                )
                st.plotly_chart(fig_ndvi_bar)

                # Gráfico de pizza
                valores = {
                    'Indicador': ['NDVI mínimo', 'NDVI máximo', 'NDVI médio'],
                    'Valor': [ndvi_min, ndvi_max, ndvi_mean]
                }
                df_pizza = pd.DataFrame(valores)
                soma = df_pizza['Valor'].sum()
                df_pizza['Percentual (%)'] = df_pizza['Valor'] / soma * 100

                fig_ndvi_pie = px.pie(
                    df_pizza,
                    names='Indicador',
                    values='Percentual (%)',
                    title='Distribuição percentual dos valores de NDVI',
                    color_discrete_sequence=[
                        "#d73027",  # vermelho (NDVI mínimo)
                        "#1a9850",  # verde (NDVI máximo)
                        "#fee08b"   # amarelo (NDVI médio)
                    ]
                )
                st.plotly_chart(fig_ndvi_pie)

                if st.button("Exportar relatório Excel (.xlsx)"):
                    caminho_arquivo = gerar_excel_estatisticas(df_stats, fig_ndvi_bar, fig_ndvi_pie)
                    with open(caminho_arquivo, 'rb') as f:
                        st.download_button(
                            label="Clique aqui para baixar o relatório",
                            data=f,
                            file_name="relatorio_ndvi.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
    else:
        st.info("Faça upload do arquivo na aba 'Dados'.")

with tab_mapa:
    if uploaded_file:
        area_fc = ler_geojson(uploaded_file)
        imagem = buscar_imagem_sentinel(area_fc)
        ndvi = calcular_ndvi(imagem)

        mapa = geemap.Map()
        mapa.add_basemap("SATELLITE")
        ndvi_params = {
            'min': 0,
            'max': 1,
            'palette': ['#d7191c', '#fdae61', '#a6d96a', '#1a9641']
        }
        mapa.addLayer(ndvi, ndvi_params, 'NDVI')
        mapa.centerObject(area_fc, 12)
        mapa.addLayer(area_fc, {}, "Área de estudo")
        mapa.to_streamlit(height=600)
    else:
        st.info("Você precisa carregar um GeoJSON na aba 'Dados'.")
